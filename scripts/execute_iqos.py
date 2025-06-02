import time
from datetime import datetime
import openpyxl
import pyautogui
from sqlalchemy import create_engine
from sqlalchemy import text
import subprocess
import os
import sys
import json
import pandas as pd


def get_resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        # base_path = root folder (one level up from current script folder)
        base_path = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    return os.path.join(base_path, relative_path)

# Use the helper function to locate cleaned_df_next.json
cleaned_df_next_path = get_resource_path('assets/cleaned_df_next.json')

# Open and load the JSON file
with open(cleaned_df_next_path, 'r') as file:
    cleaned_df_next = json.load(file)
    to_iqos = pd.DataFrame(cleaned_df_next)


print(to_iqos)

#open "Advanced Distribution" in Sage
ad_location = pyautogui.locateOnScreen(get_resource_path('assets/advanced_distribution.PNG'), confidence=0.8)
if ad_location:
    x, y, w, h = ad_location  # Get the x, y coordinates and width (w), height (h) of the image
    center_x = x + w // 2  # Calculate the center X coordinate
    center_y = y + h // 2  # Calculate the center Y coordinate

    # Print the center coordinates for verification
    print(f"Image center found at ({center_x}, {center_y})")

    # Single click at the center of the detected image
    pyautogui.click(center_x, center_y)
    print("Clicked the image!")
else:
    print("Image not found")

bit_location = pyautogui.locateOnScreen(get_resource_path('assets/branch_inventory_transfers.PNG'), confidence=0.8)
if bit_location:
    x, y, w, h = bit_location  # Get the x, y coordinates and width (w), height (h) of the image
    center_x = x + w // 2  # Calculate the center X coordinate
    center_y = y + h // 2  # Calculate the center Y coordinate

    # Print the center coordinates for verification
    print(f"Image center found at ({center_x}, {center_y})")

    # Single click at the center of the detected image
    pyautogui.click(center_x, center_y)
    print("Clicked the image!")
else:
    print("Image not found")

time.sleep(1)

# Press Down key twice and then Enter
time.sleep(1)  # Add a slight delay for navigation
pyautogui.press("down")
print("Pressed Down Arrow.")
time.sleep(1)

pyautogui.press("down")
print("Pressed Down Arrow again.")
time.sleep(1)

pyautogui.press("enter")
print("Pressed Enter.")
time.sleep(1)

pyautogui.press("enter")
print("Pressed Enter.")
time.sleep(1)

# Simulate pressing "00"
pyautogui.press("0")  # Press '0' the first time
pyautogui.press("0")  # Press '0' the second time
print("Pressed 0 twice to simulate '00'.")
time.sleep(4)

pyautogui.press("enter")
print("Pressed Enter.")
time.sleep(6)



# Check for the appearance of the message box
try:
    warning = pyautogui.locateOnScreen(get_resource_path('assets/no_transfers.PNG'), confidence=0.8)
except pyautogui.ImageNotFoundException:
    warning = None

if warning:
    print("Warning detected! Pressing Enter.")
    pyautogui.press("enter")  # Corrected indentation
    time.sleep(6)
else:
    print("No warning detected.")

pyautogui.press("escape")
print("Pressed Escape.")
time.sleep(6)  # Small delay between presses (optional)

print("No warning detected. Hitting escape.")
pyautogui.press("escape")
print("Pressed Escape.")
time.sleep(6)

print("No warning detected. Hitting escape.")
pyautogui.press("escape")
print("Pressed Escape.")
time.sleep(6)

# Process each product code, quantity, and warehouse in 'to_iqos'
for index, row in to_iqos.iterrows():
    product_code = row['product']
    quantity = row['quantity_free']
    warehouse = row['warehouse']  # Assume 'warehouse' column exists in the dataframe

    # Step 1: Enter the product code
    pyautogui.typewrite(str(product_code))  # Paste product code
    time.sleep(2)
    pyautogui.press("tab")  # Move to the quantity field
    time.sleep(2)

    # Step 2: Enter the quantity
    pyautogui.typewrite(str(quantity))  # Paste quantity
    time.sleep(2)
    pyautogui.press("f4")  # Perform F4 action
    time.sleep(2)

    # Step 3: Enter the warehouse
    pyautogui.typewrite(str(warehouse))  # Paste warehouse
    time.sleep(2)
    pyautogui.press("enter")  # Confirm warehouse entry
    time.sleep(2)

    # Step 4.1: Press the first 'Enter' to check for the warning
    pyautogui.press("enter")
    time.sleep(2)  # Allow time for the warning to appear, if applicable

    try:
        warning = pyautogui.locateOnScreen(get_resource_path('assets/reorderwarning.PNG'), confidence=0.8)
    except pyautogui.ImageNotFoundException:
        warning = None

    if warning:
        print("Warning detected! Adding five 'Enter' presses.")
        for _ in range(5):
            pyautogui.press("enter")
            time.sleep(1)
    else:
        print("No warning detected. Adding four 'Enter' presses.")
        for _ in range(4):
            pyautogui.press("enter")
            time.sleep(1)

    # Step 5: Finalize with F8
    pyautogui.press("f8")  # Perform F8 action
    time.sleep(4)

    # Load the cleaned JSON file
    with open(get_resource_path('assets/cleaned_df_next.json'), "r") as json_file:
        cleaned_data = json.load(json_file)

        # Bin Selection Process
        best_before_col_location = pyautogui.locateOnScreen(get_resource_path('assets/best_before_col.PNG'), confidence=0.8)

        if best_before_col_location:
            x, y, w, h = best_before_col_location
            center_x = x + w // 2
            center_y = y + h // 2
            pyautogui.doubleClick(center_x, center_y)
            print("Double-clicked on the 'Best Before' column!")
            time.sleep(1)

            total_bins = int(cleaned_data["total_bins_for_product"][str(index)])
            product_code = cleaned_data["product"][str(index)].strip()
            print(f"Processing {total_bins} bin(s) for product: {product_code}")

            # Make sure Sage is focused
            pyautogui.click()
            time.sleep(0.5)

            if total_bins == 1:
                # Just Alt+A once (no down)
                pyautogui.keyDown("alt")
                pyautogui.press("a")
                pyautogui.keyUp("alt")
                print(f"Selected only bin for product: {product_code}")
            else:
                # For multiple bins:
                # 1) Alt+A on first bin (pre-highlighted)
                pyautogui.keyDown("alt")
                pyautogui.press("a")
                pyautogui.keyUp("alt")
                print(f"Selected first bin for product: {product_code}")
                time.sleep(0.5)

                # 2) For each next bin (2 to total_bins):
                for bin_index in range(2, total_bins + 1):  # bin_index from 2 to total_bins
                    pyautogui.press("down")  # move down to next bin
                    time.sleep(0.5)
                    pyautogui.keyDown("alt")
                    pyautogui.press("a")
                    pyautogui.keyUp("alt")
                    print(f"Moved down and selected bin {bin_index} for product: {product_code}")
                    time.sleep(0.5)

            # Finalize bin selection
            for _ in range(4):
                pyautogui.press("enter")
                print("Pressed Enter.")
                time.sleep(1)

            pyautogui.press("escape")
            print(f"Returned to main screen for product: {product_code}")
        else:
            print("Error: 'Best Before' column not found.")

    # Optional: Add a delay if the GUI needs to load the original screen
    time.sleep(1)  # Wait briefly before moving on to the next product

    print(f"Processed Product: {product_code}, Quantity: {quantity}, Warehouse: {warehouse}")

pyautogui.press("escape")
print("Pressed Escape.")
time.sleep(3)

pyautogui.press("enter")
print("Pressed Enter.")
time.sleep(3.3)


#Generating the J number for the IQOS
# Load the workbook and select the relevant sheet
file_path = r"\\ibco-server2016\Company\Stock Control\Goods Movement Control\Weekly Movements 2020.xlsx"  # Update with your file path
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active  # Replace 'active' with the specific sheet name if needed

# Find the earliest blank row in the "Date" column (Column 2)
next_row = None
for row in range(6, sheet.max_row + 1):  # Start from row 6 since rows 1–5 are headers
    date_cell = sheet.cell(row=row, column=2)  # Column 2 corresponds to the "Date" column
    if date_cell.value is None:  # If the "Date" column is blank
        next_row = row
        break

# Ensure a blank row is found
if next_row is not None:
    # Retrieve the existing J reference from Column A (Column 1)
    j_reference = sheet.cell(row=next_row, column=1).value  # Get the existing J reference
    if j_reference is None:
        print(f"No J reference found in row {next_row}.")
    else:
        # Fill in the remaining details in the same row
        sheet.cell(row=next_row, column=2, value=datetime.now().strftime("%d/%m/%Y"))  # Set the current date in column "Date"
        sheet.cell(row=next_row, column=4, value="IQOS Transfer")  # Set "Type"
        sheet.cell(row=next_row, column=6, value="MIX PRODUCTS")  # Set "Product"
        sheet.cell(row=next_row, column=9, value="WH96/01")  # Set "Goods In"
        sheet.cell(row=next_row, column=10, value="WH00")  # Set "Goods Out"

        # Save the changes to the workbook
        workbook.save(file_path)

        print(f"Entry added successfully in row {next_row}:")
        print(f"Ref No: {j_reference}, Date: {datetime.now().strftime('%d/%m/%Y')}, "
              f"Type: IQOS Transfer, Product: MIX PRODUCTS, Goods In: WH96/01, Goods Out: WH00")
else:
    print("No available row found for entry.")


# Now use the J reference in Sage
time.sleep(2)  # Ensure the Sage interface is ready

# Type the J reference into the relevant field
pyautogui.typewrite(j_reference)  # Enter the J reference
print(f"Entered J reference: {j_reference}")
time.sleep(2)

# Submit or proceed as needed
pyautogui.press("enter")
print("Submitted the J reference.")
time.sleep(1)

# Submit or proceed as needed
pyautogui.press("enter")
print("Generated Sage Print.")
time.sleep(4)

#Discard the print from Sage
pyautogui.press("f5")  # Perform F5 action
time.sleep(2)

pyautogui.press("enter")
print("Confirmed IQOS.")
time.sleep(2)

pyautogui.press("escape")
print("Pressed Escape.")
time.sleep(3)

pyautogui.press("enter")
print("Pressed Enter.")
time.sleep(3.3)

pyautogui.press("escape")
print("Pressed Escape - should be back to main screen.")
time.sleep(6)

#transfernoteprint
pyautogui.press("down")
print("Pressed down.")
time.sleep(2)  # Add a slight delay for reliability


pyautogui.press("enter")
print("Pressed Enter.")
time.sleep(2)

pyautogui.press("enter")
print("Pressed Enter.")
time.sleep(2)

pyautogui.press("i")
print("Pressed i for individual.")
time.sleep(2)

pyautogui.press("enter")
print("Pressed Enter.")
time.sleep(2)


# Load the configuration file
with open('assets/db_config.json', 'r') as config_file:
    config = json.load(config_file)




# Construct the SQLAlchemy connection string
conn_str = (
    f"mssql+pyodbc://{config['database']['username']}:{config['database']['password']}@"
    f"{config['database']['server']}/{config['database']['database_name']}?"
    f"driver=ODBC+Driver+17+for+SQL+Server"
)

# Create an engine using SQLAlchemy
engine = create_engine(conn_str)

# Create a connection using the engine
with engine.connect() as conn:
    # Define the SQL query using SQLAlchemy's `text` object
    query = text("""
    SELECT [key_value]
    FROM [ibco].[scheme].[sysdirm]
    WHERE [system_key] = 'AWTRNLAST'
    """)

    # Execute the query and fetch the result
    result = conn.execute(query)
    key_value = result.scalar()  # Fetch the first column of the first row
    prefixed_value = f"IBC{key_value}"  # Prefix the value with "IBC"
    print(f"Prefixed Value: {prefixed_value}")

    # Automate entry with pyautogui
    import pyautogui
    import time
    time.sleep(2)
    pyautogui.typewrite(prefixed_value)
    pyautogui.press("enter")
    print("Entered the value into the application.")


pyautogui.press("enter")
print("Pressed Enter.")
time.sleep(2)

pyautogui.press("enter")
print("Pressed Enter.")
time.sleep(2)

pyautogui.press("enter")
print("Pressed Enter.")
time.sleep(2)

#Allocate picking list number
pyautogui.press("f6")  # Perform F6 action
time.sleep(2)

pyautogui.press("enter")
print("Pressed Enter.")
time.sleep(2)

pyautogui.press("enter")
print("Pressed Enter.")
time.sleep(5)

#Allocate picking list number
pyautogui.press("f5")  # Perform F5 action
time.sleep(2)

pyautogui.press("enter")
print("Pressed Enter.")
time.sleep(5)

#Discard print in Sage
pyautogui.press("f5")  # Perform F5 action
time.sleep(2)

pyautogui.press("escape")
print("Pressed Escape.")
time.sleep(3)


# Autogenerating the Crystal Report for the User.

# Define the path to the Crystal Reports executable and the .rpt file
crystal_exe = r"C:\PROGRA~2\SaberLogic\Logicity\Logicity Desktop.exe"  # Path to Logicity Desktop
rpt_file_path = get_resource_path("assets/iqos_auto_print.rpt")  # Path to the Crystal Report file

# Check if the .rpt file exists in the same directory
if not os.path.exists(rpt_file_path):
    print(f"Error: File {rpt_file_path} not found.")
else:
    # Execute the command with subprocess
    try:
        subprocess.run([crystal_exe, rpt_file_path], check=True)
        print(f"Crystal Report file '{rpt_file_path}' opened successfully.")
    except FileNotFoundError:
        print(f"Error: Executable '{crystal_exe}' not found.")
    except subprocess.CalledProcessError as e:
        print(f"Error: {e}")


# Close the connection
conn.close()





